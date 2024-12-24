# CREATED: 21-DEC-2024
# LAST EDIT: 24-DEC-2024
# AUTHOR: DUANE RINEHART, MBA (duane.rinehart@gmail.com)

'''
GOALS:
1) ANALYZE HISTORICAL STOCK INFORMATION, CURRENT OPTIONS PREMIUMS (COVERED CALLS) TO
CALCULATE EXPECTED VALUE OF OPTIONS CONTRACTS (AND COMPARE TO CURRENT PRICES)

2) CALCULATE ANNUALIZED RETURN ON [COVERED CALLS] OPTIONS CONTRACTS


INPUTS:
1) EXCEL FILE OF STOCKS OF INTEREST (WHERE >= 100 SHARES ARE OWNED)
2) API TO FINANCIAL INFORMAITON SERVICE (E.G. YAHOO FINANCE) FOR HISTORICAL STOCK INFORMATION

OUTPUTS:
2) EXCEL DOCUMENT CONTAINING ANALYSIS OF STOCK LIST (CALCS)

INTERMEDIATE DATA STORE:
SQL LITE

FILE & EXCEL LOCATIONS DEFINED IN constants.json.  See setup.ipynb for creation


'''

import pandas as pd
import numpy as np
import json
import ast
from data_access import YahooAPI
from dashboard import populate_db, db_table_stats
from data_access import HistData
from pathlib import Path
import sqlite3
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from datetime import datetime


def load_app_constants():
    # Load constants from constants.json
    constants_file_path = 'constants.json'

    try:
        with open(constants_file_path, 'r') as file:
            content = file.read()
            # Adjust the content to be compatible with Python syntax
            content = content.replace("constants = {", "{").replace("};", "}")
            # Use ast.literal_eval to safely evaluate the string as a Python dictionary
            constants = ast.literal_eval(content)
            return constants
    except FileNotFoundError:
        print(f"File {constants_file_path} not found.  Did you run setup.ipynb?")
        return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None


def db_last_update(dict_constants):
    db_server = dict_constants.get('HOST_DB_SERVER', 'sqlite')
    db_name = dict_constants.get('HOST_DB_NAME', 'finance.db')

    results = db_table_stats(db_server, db_name, 'history')
    print('TICKER', 'START_DATE', 'END_DATE', sep='\t')
    print('------', '----------', '--------', sep='\t')
    [print(row[0], row[1], row[2], sep='\t') for row in results]


def db_update(dict_constants):
    '''
    populate_db assume table to be populated is 'history', consistent with db_schema.sql
    '''
    print("Updating database...")
    db_server = dict_constants.get('HOST_DB_SERVER', 'sqlite')
    db_name = dict_constants.get('HOST_DB_NAME', 'finance.db')
    input_excel_file_path = dict_constants.get('STOCK_EXCEL_INPUT_FILE_PATH', '')
    input_excel_file_name = dict_constants.get('STOCK_EXCEL_INPUT_FILE_NAME', 'stock_blotter.xlsx')
    input_excel_stocks = Path(input_excel_file_path, input_excel_file_name)
    
    # Read the Excel file into a DataFrame (from 'inputs' worksheet)
    df = pd.read_excel(input_excel_stocks, sheet_name='inputs')
    symbols = df['SYMBOL'].tolist() #EXTRACT JUST SYMBOLS

    populate_db(db_server, db_name, 'history', symbols, '1y')


def create_output_analysis(dict_constants):
    db_server = dict_constants.get('HOST_DB_SERVER', 'sqlite')
    db_name = dict_constants.get('HOST_DB_NAME', 'finance.db')
    input_excel_file_path = dict_constants.get('STOCK_EXCEL_INPUT_FILE_PATH', '')
    input_excel_file_name = dict_constants.get('STOCK_EXCEL_INPUT_FILE_NAME', 'stock_blotter.xlsx')
    input_excel_stocks = Path(input_excel_file_path, input_excel_file_name)
    output_excel_file_path = Path(dict_constants.get('STOCK_EXCEL_OUTPUT_FILE_PATH', 'output'))
    output_excel_file_name = dict_constants.get('STOCK_EXCEL_OUTPUT_FILE_NAME', 'analysis_output.xlsx')
    output_excel_analysis = Path(output_excel_file_path, output_excel_file_name)

    # Read the Excel file into a DataFrame (from 'inputs' worksheet)
    df_input = pd.read_excel(input_excel_stocks, sheet_name='inputs')
    symbols = df_input['SYMBOL'].tolist() #EXTRACT JUST SYMBOLS
    sql_where = "('" + "', '".join(symbols) + "')"
    conn = sqlite3.connect(db_name)  # Replace with your DB connection details
    sql = f'SELECT * FROM history WHERE symbol IN {sql_where}'
    df = pd.read_sql(sql, conn) 
    conn.close()

    # Sort by 'symbol' (ascending) and 'date' (descending); remove time from datetime stamp
    df_historical = df.sort_values(by=['symbol', 'date'], ascending=[True, False])
    df_historical['date'] = pd.to_datetime(df_historical['date'], errors='coerce')
    df_historical['date'] = df_historical['date'].dt.date

    #CALCS FOR ANLALYSIS OUTPUT
    for symbol in symbols:
        oldest_close_price, newest_close_price, days_difference, price_target, std_dev = calc_price_target(df, symbol)
        
    df_analysis = df_input[['SYMBOL']].copy()
    df_analysis[['oldest_close_price', 'newest_close_price', 'price diff period', 'price_target', 'std_dev (cur month)']] = df_input.apply(lambda row: pd.Series(calc_price_target(df, row['SYMBOL'])), axis=1)
    df_analysis['strike price (+1mo)'] = None
    df_analysis['EV (min premium)'] = None

    output_excel_file_path.mkdir(parents=True, exist_ok=True)
    try:
        with pd.ExcelWriter(output_excel_analysis, mode='w', engine='openpyxl') as writer:
            df_historical.to_excel(writer, index=False, sheet_name='historical')
            df_analysis.to_excel(writer, index=False, sheet_name='analysis')

            #Format column widths to fit text (on analysis sheet)
            wb = writer.book
            ws = wb['analysis']

            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Apply formulas to 'EV' column (starting from row 2 to skip the header)
            for row in range(2, len(df_analysis) + 2):  # Start from row 2 (skip header)
                ws[f'F{row}'] = f"=C{row}+((C{row}-B{row})/D{row}*30)"  # Set the price target formula
                ws[f'H{row}'] = f"=F{row}/SQRT(6.28)*EXP(-1/2*POWER((G{row}-E{row})/F{row},2)) - (G{row} - E{row})/2*(1-ERF((G{row}-E{row})/SQRT(2*F{row}*F{row})))"  # Set the EV formula
                ws[f'G{row}'] = f"=ROUND(C{row},0)+ROUNDUP(C{row}*1.1/12, 0)" # Set the price target formula (10% annualized gain)
                ws[f'G{row}'].fill = yellow_fill # Highlight column G in yellow
            
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

            ws.column_dimensions['H'].width = 32 #force EV column width

        print(f"Data exported successfully to '{output_excel_analysis}'")
    except PermissionError as e:
        print(f"PermissionError: {e}. Please check file permissions and ensure the file is not open in another program.")
    except FileNotFoundError as e:
        print(f"FileNotFoundError: {e}. The specified directory or file does not exist.")
    except Exception as e:
        print(f"An error occurred: {e}. Please check the file path and ensure all necessary libraries are installed.")
        


def calc_price_target(df, symbol):
    df['date'] = pd.to_datetime(df['date'])
    # Filter the DataFrame for individual symbol
    individual_stock_df = df[df['symbol'] == symbol]

    oldest_date = individual_stock_df['date'].min()
    newest_date = individual_stock_df['date'].max()

    # Get the close prices for these dates
    oldest_close_price = individual_stock_df.loc[individual_stock_df['date'] == oldest_date, 'close'].iloc[0]
    newest_close_price = individual_stock_df.loc[individual_stock_df['date'] == newest_date, 'close'].iloc[0]

    days_difference = (newest_date - oldest_date).days

    # Calculate the 30-day price target [replaced with Excel formula]
    price_change_over_period = newest_close_price - oldest_close_price
    avg_daily_price_change = price_change_over_period / days_difference
    price_target = newest_close_price + (avg_daily_price_change * 30)

    # Filter the dataframe to only include the most recent month's data
    current_month = datetime.now().month
    current_year = datetime.now().year

    recent_data = individual_stock_df[
        (individual_stock_df['date'].dt.month == current_month) & 
        (individual_stock_df['date'].dt.year == current_year)
    ]

    # Calculate the standard deviation of the 'close' column
    std_dev = recent_data['close'].std()

    return oldest_close_price, newest_close_price, days_difference, price_target, std_dev
 

def calc_monthly_standard_deviation(stock):
    # Example data
    data = pd.read_csv('stock_data.csv')
    data['Returns'] = data['Close'].pct_change()
    std_dev = data['Returns'].std()
    monthly_std_dev = std_dev * np.sqrt(21)  # Approx. 21 trading days per month
    print(monthly_std_dev)


'''
constants = {
    "stock_symbol": "AAPL",
    "strike_price": 150,
    "option_premium": 2.5,
    "risk_free_rate": 0.05,
    "time_to_expiration_days": 30,
    "volatility_estimate": 0.2
}
'''
def display_menu():
    '''ENSURE CHOICE OPTIONS EXIST FOR EACH ITEM'''
    print('1 = UPDATE DATABASE')
    print('2 = SHOW LAST DATABASE INFO')
    print('3 = ANALYZE OPTIONS')
    print('M = DISPLAY MENU')
    print('X = EXIT')


def main():
    print('WELCOME TO OPTIONS ANALYSIS TRACKER')
    print('------------------------------------')
    dict_constants = load_app_constants()
    display_menu()

    while True:
        choice = input('ACTION: ')
        if choice == '1':
            db_update(dict_constants)
        if choice == '2':
            db_last_update(dict_constants)
        elif choice == '3':
            print('ANALYZING OPTIONS...')
            create_output_analysis(dict_constants)
        elif choice.upper() == 'M':
            display_menu()
        elif choice.upper() == 'X':
            print('GOODBYE')
            break
        else:
            print('INVALID CHOICE')


if __name__ == "__main__":
    main()